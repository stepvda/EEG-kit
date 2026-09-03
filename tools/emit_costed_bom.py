#!/usr/bin/env python3
"""
emit_costed_bom.py -- the programme's INTERNAL costed bill of materials.

  docs/EEG_kit_BOM_INTERNAL_RevC_costed.xlsx

Package v1 carried a real internal cost model: an Assumptions sheet, a BOM priced at four
fleet quantities, and a roll-up by subsystem.  Package v2 Rev B lost all of it -- the file
kept the name and became a copy of the bidders' workbook with two empty price columns.
This script rebuilds the model on v2's part list.  See ECO-EEG-016 section 2A, C-14.

Prices are INDICATIVE ESTIMATES in EUR, ex VAT and ex shipping, exactly as v1's were.  Each
line records the basis it was estimated from, so a reviewer can challenge a number rather
than having to take the sheet on trust.  The manufacturer's quotation replaces them.

The line list comes from emit_workbooks.KIT, so this workbook and the bidders' workbook can
never disagree about what is in a kit.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import emit_workbooks as W

HERE = os.path.dirname(os.path.abspath(__file__))
DOCS = os.path.join(os.path.dirname(HERE), "docs")

REV = "C"
DATE = "1 September 2026"

# --------------------------------------------------------------------------- assumptions
SCENARIOS = [("A -- prototypes", 2), ("B -- pilot fleet", 10),
             ("C -- fleet, lower bound", 25), ("D -- fleet, upper bound", 50)]
SPARE_SHARE = 0.25
EUR_PER_USD = 0.92
NRE = [("Layout review NRE placeholder (EUR)", 1500,
        "RFQ-EEG-002A is a REVIEW of supplied routing, not a layout job. v1 budgeted "
        "EUR 1500 for the layout itself; the review is the cheaper half of that range, "
        "but the figure is held until the quote."),
       ("Foam insert tooling placeholder (EUR)", 250,
        "One-off die or laser programme. HELD AT v1'S FIGURE AND KNOWN TO BE LOW: the "
        "CASE-00 Rev C stack is SEVEN layers, not one or two, and the layers do not share "
        "a profile -- the bays cut through different layer counts, and layer 1 of the "
        "helmet bay is 8 mm smaller all round for the halo shelf -- so this is several "
        "distinct programmes, not one. The figure is not re-estimated here because the "
        "Rev C cut file has not been drawn (PKG-EEG-015 open item 1)."),
       ("Stencil, fixtures and test jig placeholder (EUR)", 400,
        "JIG-EEG-009 Rev B specifies more fixtures than v1 assumed. Held at v1's figure "
        "until JIG-EEG-009 section 6.1 is quoted.")]

# ------------------------------------------------------------------- the price table
# item number (1-based, matching emit_workbooks.KIT order) -> (@2, @10, @25, @50, basis)
# Every figure is an indicative estimate.  "v1 line N" means the price is carried from
# package v1's EEG_kit_BOM_INTERNAL_costed.xlsx, sheet BOM, at the same four breaks.
COST = {
    # ---- Modules.  v2 buys modules where v1 bought bare silicon (DSN-EEG-003 section 2),
    #      so these lines are all dearer than their v1 ancestors and that is the point of
    #      the architecture: no fine-pitch rework, no BGA, no ADS1299 lead time.
    1:  (95, 82, 74, 68, "v1 line 3 was the bare ADS1299 at 55/48/45/42; a breakout adds "
                         "its own PCB, regulators, crystal and headers. Two per kit."),
    2:  (16, 14, 12.5, 11.5, "DevKitC-1-N16R8, not the bare WROOM of v1 line 15 (4.20 at "
                             "the 2 break). The kit uses the DevKit's own USB and UART."),
    3:  (9, 7.5, 6.5, 6, "ES8388 breakout. v1 line 11 was the bare codec at 1.60."),
    4:  (22, 18, 16, 15, "v1 lines 19+20 (ADuM4160 9.50 + isolated DC-DC 3.50) on a "
                         "carrier board. E-24 requires the isolator from Rev A."),
    5:  (7, 6, 5.5, 5, "Adafruit ATECC608 breakout 4314. v1 line 16 was the bare part."),
    6:  (16, 14, 12.5, 11.5, "Adafruit bq24074 4755. v1 line 26 was the bare BQ24075."),
    7:  (11, 9.5, 8.5, 8, "Adafruit MAX17048 5580. v1 line 27 was the bare gauge."),
    8:  (13, 11, 9.5, 8.5, "TPS63020 breakout. NEW IN v2: package v1 had no source for "
                           "the 3.3 V rail at all (ECO-EEG-016, blocking finding)."),
    9:  (4, 3, 2.6, 2.3, "microSD SDMMC breakout. v1 line 17 was the bare socket."),
    10: (8, 6.5, 5.8, 5.2, "MAX4466-class preamp module, Adafruit 1063 class."),
    11: (9, 7.5, 6.5, 6, "OPEN LINE: the room-microphone module is not selected "
                         "(kit BOM note, AVL-EEG-017). Priced as a MAX4466-class board "
                         "plus capsule; the figure moves when the part is chosen."),
    12: (4.5, 3.5, 3, 2.6, "74HC595 breakout. NEW IN v2: in v1 the eight contact-light "
                           "lines had nothing driving them (ECO-EEG-016)."),
    13: (8, 7, 6.5, 6, "v1 line 24, unchanged."),
    14: (9, 8, 7.5, 7, "v1 line 18, unchanged."),

    # ---- Carrier
    15: (30, 16, 11, 8, "v1 line 1 was 12/8/6/5 for a ~95x65 two-layer board. v2's is "
                        "150.0 x 130.0 mm and FOUR layers. README_package_index.txt puts "
                        "the four-layer premium at about EUR 35 in total at 2 units and "
                        "about EUR 3 a board at 50; this line carries that."),
    16: (75, 45, 32, 25, "v1 line 2 was 60/35/25/20 for 112 placements. v2 has 189 "
                         "(153 SMT placed parts + 33 through-hole + 3 fiducials)."),
    17: (7, 5.5, 4.5, 4, "Made-up 2.54 mm keyed IDC ribbon set, WH-EEG-008 / "
                         "ICD-EEG-006 section 3. NEW IN v2 with the module architecture."),

    # ---- Printed.  MJF PA12 at a bureau, priced by volume; see mech/MANIFEST.json.
    18: (45, 34, 27, 22, "HM-01, 133.6 cm3, the largest printed part in the kit."),
    19: (2.6, 2.0, 1.6, 1.4, "HM-04, 1.9 cm3 each, TEN per kit (8 fitted + 2 spare)."),
    20: (4, 3, 2.5, 2.2, "HM-08, 6.9 cm3. Called HM-07 in v1; renumbered because HM-07 "
                         "named two different parts (PARTS-EEG-019, ECO-EEG-015)."),
    21: (6, 4.5, 3.8, 3.2, "HM-07 boom arm, printed mount only; the 120 mm gooseneck is "
                           "line 29 and the capsule is line 40. v1 line 14 bundled all "
                           "three at 14/11/9/8."),
    22: (12, 9, 7.5, 6.5, "MP-01, the module mounting plate. NEW IN v2."),
    23: (46, 34, 27, 22, "v1 line 34 was 38/28/22/18 for a 146.8 x 140.8 x 44 mm pod. "
                         "POD-P1 is now 163 x 143 x 58 mm to take the bigger carrier."),
    24: (8, 6.5, 5.5, 5, "HM-02, TPU 85A comfort pads x4 plus the chin-cup liner."),
    25: (3, 2.2, 1.8, 1.5, "HM-09 service key. Qty 0 per kit -- one per SITE, not per "
                           "kit, so it contributes nothing to the per-kit cost."),
    26: (1.5, 1.2, 1, 0.9, "FIT-01 coupon. Qty 0 per kit; printed once per build batch."),

    # ---- Bought-in mechanical
    27: (0.35, 0.28, 0.24, 0.21, "Stainless compression springs, 3-6 N, ten per kit."),
    28: (4.5, 3.5, 3, 2.6, "Hard-hat style occipital ratchet with POM pawl."),
    29: (3.5, 2.8, 2.4, 2.1, "120 mm gooseneck and mount; part of v1's bundled line 14."),
    30: (4, 3.2, 2.7, 2.4, "Chin strap webbing, cup and anchors."),
    31: (0.4, 0.3, 0.26, 0.22, "M3 x 18 nylon standoffs and M3 x 6 screws, four sets."),
    32: (18, 14, 11.5, 10, "WH-01, 12-way screened custom assembly, WH-EEG-008. v1 had "
                           "one 20-way harness; ECO-EEG-014 split it into two cables."),
    33: (12, 9, 7.5, 6.5, "WH-02, 10-way contact-light harness. NEW IN v2 -- it did not "
                          "exist in v1 because the lights had no driver."),

    # ---- Electrodes
    34: (9.5, 8.5, 7.8, 7.5, "v1 line 37. Greentek MOQ is 200 pieces, which covers about "
                             "16 kits; AVL-EEG-017 section 4 governs the order, not the "
                             "build. TEN per kit on the service bayonet."),
    35: (8, 7, 6.5, 6, "v1 line 38, unchanged."),
    36: (4, 3.5, 3.2, 3, "v1 line 39, unchanged."),
    37: (9, 8, 7.5, 7, "v1 line 40, unchanged. Consumable, replaced at refurbishment."),
    38: (0.25, 0.2, 0.16, 0.14, "Two-lead bicolour red/green 3 mm, ten per kit. NEW IN "
                                "v2: v1 specified single-colour status LEDs (line 23)."),

    # ---- Audio
    39: (45, 42, 40, 38, "v1 line 42, unchanged. THIS LINE IS CITED BY SVC-EEG-013 "
                         "section 7.4: the EUR 40 charged for missing headphones is this "
                         "line at the 25 break."),
    40: (9, 7.5, 6.5, 6, "Primo EM272-class electret capsule and windscreen."),

    # ---- Consumables, cables, case, service
    41: (29.5, 25.2, 23, 21.9, "v1 lines 43+44+45+46 summed at their own quantities: "
                               "paste 2 x 7.00, prep gel 8.00, wipes/gauze set 6.00, "
                               "pouch 1.50 at the 2 break."),
    42: (5, 4, 3.6, 3.2, "v1 lines 47+48, both cables."),
    43: (6, 5, 4.5, 4, "v1 line 49, unchanged."),
    44: (250, 225, 205, 190, "Rev B priced 78/66/54/46 against a shell class "
                             "(Nanuk 915 / Peli 1450) that is now withdrawn, and v1 line "
                             "50 priced 65/55/45/38 for a ~300 x 220 x 110 case before "
                             "that. PKG-EEG-015 "
                             "section 3.2 settles the envelope at >= 516 x 390 x 210 mm "
                             "and the shell at the Peli 1560, unwheeled and no-foam, which "
                             "is a large-format case several classes up, not a size up. "
                             "Indicative single-distributor list less trade discount at "
                             "the four breaks. No shell has been bought, measured or "
                             "quoted, so this line is the least tested in the sheet."),
    45: (282, 207, 158, 124, "Rev B's 34/25/19/15 priced two 340 x 250 x 25 mm sheets, "
                             "4.25 litres. The CASE-00 Rev C stack of PKG-EEG-015 section "
                             "2.2 is seven 516 x 390 x 25 mm layers, 35.2 litres: 8.29x "
                             "the material and 8.29x the cut length, and the Rev B figures "
                             "are scaled by that ratio at each break. Material and cutting "
                             "only; the die or laser programme is the one-off on the "
                             "Assumptions sheet. Not quoted, and the Rev C cut file has "
                             "not been drawn."),
    46: (5, 4, 3.2, 2.8, "v1 lines 52+53, the laminated card and the shipping carton."),
    47: (45, 25, 18, 15, "v1 line 54, unchanged. TST-EEG-004 Rev C is 31 steps; the "
                         "figure is held until a unit has actually been through them."),
}

ELECTRONICS_GROUPS = {"Modules", "Carrier"}

# package v1's own per-kit totals, from EEG_kit_BOM_INTERNAL_costed.xlsx sheet
# "BOM", the cached values of =SUM(M2:M55) .. =SUM(P2:P55).  Carried so that the
# comparison in the Kit summary sheet is against a real prior figure, not a memory.
V1_PER_KIT = [740.13, 598.60, 523.32, 474.95]

# ------------------------------------------------------------------------------- style
HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
INPUT = PatternFill("solid", fgColor="FFF2CC")
TOTAL = PatternFill("solid", fgColor="D9E2F3")
NOTE = Font(italic=True, size=9, color="444444")
BOLD = Font(bold=True, size=10)
TITLE = Font(bold=True, size=13, color="1F3864")
BOX = Border(*[Side(style="thin", color="BFBFBF")] * 4)
EUR = '#,##0.00'


def _hdr(ws, row=1):
    for c in ws[row]:
        if c.value is not None:
            c.fill, c.font = HDR_FILL, HDR_FONT
            c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


# --------------------------------------------------------------------------- build
def build(path):
    wb = openpyxl.Workbook()

    # ============================================================ sheet 1: Assumptions
    ws = wb.active
    ws.title = "Assumptions"
    ws.append(["EEG field kit -- internal bill of materials, indicative costing"])
    ws["A1"].font = TITLE
    ws.append([f"Rev {REV}, {DATE}.  Companion to RFQ-EEG-001 Rev E and DSN-EEG-003 Rev C."])
    ws.append(["THIS IS THE PROGRAMME'S OWN COPY AND IT CARRIES PRICES.  The bidders' "
               "workbook (EEG_kit_BOM_for_bidders_RevC.xlsx) is the same part list with "
               "the prices removed and empty cells for the bidder.  Do not send this file "
               "to a bidder."])
    ws["A3"].font = NOTE
    ws.append(["All unit prices are indicative single-source estimates in EUR, ex VAT and "
               "ex shipping, and are placeholders for the manufacturer's quotation. Every "
               "line records the basis it was estimated from, in the BOM sheet's last "
               "column. Yellow cells are inputs; edit those and everything recalculates."])
    ws["A4"].font = NOTE
    ws.append([])
    ws.append(["Assumption", "Value", "Note"])
    _hdr(ws, ws.max_row)
    ws.append(["Fleet quantity scenarios (units)", None,
               "Quantity breaks the BOM is priced at; RFQ-EEG-001 Rev E section 10 asks "
               "bidders for these four."])
    for name, n in SCENARIOS:
        ws.append([f"Scenario {name}", n])
        ws.cell(row=ws.max_row, column=2).fill = INPUT
    ws.append(["Spare assembled boards (share of fleet)", SPARE_SHARE,
               "Build 25 % spares of hand-assembled units. Survives in RFQ-EEG-001 Rev E, "
               "Phase 3."])
    ws.cell(row=ws.max_row, column=2).fill = INPUT
    ws.append(["EUR per USD", EUR_PER_USD,
               "Used only where a supplier quotes in USD. Update on the day of the quote."])
    ws.cell(row=ws.max_row, column=2).fill = INPUT
    for label, val, note in NRE:
        ws.append([label, val, note])
        ws.cell(row=ws.max_row, column=2).fill = INPUT
    ws.append([])
    ws.append(["Legend"])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.append(["Yellow fill = input you can edit.  Plain = formula.  Prices are per unit "
               "at the stated quantity break; the manufacturer's quote replaces them."])
    ws.cell(row=ws.max_row, column=1).font = NOTE
    ws.append([])
    ws.append(["Nothing in this package has been manufactured or measured, and no price "
               "here has been quoted by a supplier. Every figure is an estimate."])
    ws.cell(row=ws.max_row, column=1).font = NOTE
    _widths(ws, [44, 12, 96])
    for row in ws.iter_rows():
        for c in row:
            c.alignment = Alignment(vertical="top",
                                    wrap_text=(c.column == 3 or c.row in (3, 4)))

    # =================================================================== sheet 2: BOM
    ws = wb.create_sheet("BOM")
    breaks = [n for _, n in SCENARIOS]
    cols = (["Item", "Group", "Description", "Reference part / module", "Qty per kit"]
            + [f"Unit EUR @{n}" for n in breaks]
            + [f"Per-kit EUR @{n}" for n in breaks]
            + ["Substitutable", "Price basis"])
    ws.append(cols)
    _hdr(ws)

    missing = [i for i in range(1, len(W.KIT) + 1) if i not in COST]
    if missing:
        raise SystemExit(f"no price for kit line(s) {missing} -- COST is out of step "
                         f"with emit_workbooks.KIT")
    if len(COST) != len(W.KIT):
        raise SystemExit(f"COST has {len(COST)} lines, KIT has {len(W.KIT)}")

    first = 2
    for i, (grp, desc, ref, _alt, qty, sub, _note) in enumerate(W.KIT, start=1):
        p2, p10, p25, p50, basis = COST[i]
        r = ws.max_row + 1
        ws.append([i, grp, desc, ref, qty, p2, p10, p25, p50,
                   f"=$E{r}*F{r}", f"=$E{r}*G{r}", f"=$E{r}*H{r}", f"=$E{r}*I{r}",
                   sub, basis])
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BOX
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (3, 4, 15)))
            if 6 <= c <= 9:
                cell.fill, cell.number_format = INPUT, EUR
            elif 10 <= c <= 13:
                cell.number_format = EUR
    last = ws.max_row

    ws.append([])
    r = ws.max_row + 1
    ws.append([None, None, "Total per complete kit (materials + assembly + test)",
               None, None, None, None, None, None]
              + [f"=SUM({col}{first}:{col}{last})" for col in "JKLM"])
    for c in range(1, len(cols) + 1):
        ws.cell(row=r, column=c).fill = TOTAL
        ws.cell(row=r, column=c).font = BOLD
        if 10 <= c <= 13:
            ws.cell(row=r, column=c).number_format = EUR
    ws.append([])
    ws.append([None, None,
               "Prices are indicative estimates for the manufacturer's quote to replace. "
               "Unit columns are per-unit prices at 2 / 10 / 25 / 50 kits; per-kit columns "
               "multiply by the quantity in this kit. Lines 25 and 26 are qty 0 per kit "
               "(HM-09 is one per site, FIT-01 one per build batch), so they cost nothing "
               "per kit and are priced here only so the figure exists when it is ordered."])
    ws.cell(row=ws.max_row, column=3).font = NOTE
    ws.cell(row=ws.max_row, column=3).alignment = Alignment(wrap_text=True, vertical="top")
    _widths(ws, [6, 15, 54, 34, 9] + [12] * 8 + [13, 70])

    # =========================================================== sheet 3: Kit summary
    ws = wb.create_sheet("Kit summary")
    ws.append(["Cost roll-up by subsystem and scenario"])
    ws["A1"].font = TITLE
    ws.append([f"Every cell is a formula over the BOM sheet, rows {first} to {last}. "
               f"Change a price there and this sheet follows."])
    ws["A2"].font = NOTE
    ws.append([])
    ws.append(["Subsystem"] + [f"Per kit @{n}" for n in breaks])
    _hdr(ws, ws.max_row)

    groups = []
    for grp, *_ in W.KIT:
        if grp not in groups:
            groups.append(grp)
    for grp in groups:
        r = ws.max_row + 1
        ws.append([grp] + [f"=SUMIF(BOM!$B${first}:$B${last},$A{r},BOM!{c}${first}:{c}${last})"
                           for c in "JKLM"])
        for c in range(2, 6):
            ws.cell(row=r, column=c).number_format = EUR
    gfirst, glast = ws.max_row - len(groups) + 1, ws.max_row

    r = ws.max_row + 1
    ws.append(["Total per kit"] + [f"=SUM({c}{gfirst}:{c}{glast})" for c in "BCDE"])
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = TOTAL
        ws.cell(row=r, column=c).font = BOLD
        if c > 1:
            ws.cell(row=r, column=c).number_format = EUR
    total_row = r

    r = ws.max_row + 1
    elec = "+".join(f"{{c}}{gfirst + groups.index(g)}" for g in groups
                    if g in ELECTRONICS_GROUPS)
    ws.append(["Electronics only (Modules + Carrier)"]
              + [f"={elec.format(c=c)}" for c in "BCDE"])
    for c in range(2, 6):
        ws.cell(row=r, column=c).number_format = EUR
    ws.cell(row=r, column=1).font = NOTE

    ws.append([])
    ws.append(["Programme build-up by scenario"])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.append(["Line"] + [f"@{n}" for n in breaks])
    _hdr(ws, ws.max_row)
    qrow = ws.max_row + 1
    ws.append(["Kits built"] + breaks)
    for c in range(2, 6):
        ws.cell(row=qrow, column=c).fill = INPUT
    r = ws.max_row + 1
    ws.append(["Kit cost x quantity"]
              + [f"={c}{total_row}*{c}{qrow}" for c in "BCDE"])
    kitcost = r
    r = ws.max_row + 1
    ws.append([f"Spare assembled carriers ({SPARE_SHARE:.0%} of fleet, carrier lines only)"]
              + [f"=ROUND({c}{qrow}*{SPARE_SHARE},0)*"
                 f"SUMIF(BOM!$B${first}:$B${last},\"Carrier\",BOM!{d}${first}:{d}${last})"
                 for c, d in zip("BCDE", "JKLM")])
    spares = r
    nre_rows = []
    for label, val, _ in NRE:
        r = ws.max_row + 1
        ws.append([label.replace(" placeholder (EUR)", "")] + [val] * 4)
        for c in range(2, 6):
            ws.cell(row=r, column=c).fill = INPUT
        nre_rows.append(r)
    r = ws.max_row + 1
    ws.append(["Programme total"]
              + [f"={c}{kitcost}+{c}{spares}+" + "+".join(f"{c}{n}" for n in nre_rows)
                 for c in "BCDE"])
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = TOTAL
        ws.cell(row=r, column=c).font = BOLD
    for row in ws.iter_rows(min_row=4):
        for c in row:
            if c.column > 1 and isinstance(c.value, str) and c.value.startswith("="):
                c.number_format = EUR
    ws.append([])
    ws.append(["The one-off lines are placeholders from the Assumptions sheet, not quotes. "
               "The spare-carrier line prices spare BOARDS, not spare kits: a spare kit "
               "would need electrodes, a case and a helmet as well."])
    ws.cell(row=ws.max_row, column=1).font = NOTE

    # ---- what package v1 estimated, and why v2 costs more
    ws.append([])
    ws.append(["What package v1 estimated, for comparison"])
    ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.append(["Line"] + [f"@{n}" for n in breaks])
    _hdr(ws, ws.max_row)
    v1row = ws.max_row + 1
    ws.append(["v1 per-kit total (EEG_kit_BOM_INTERNAL_costed.xlsx)"] + V1_PER_KIT)
    r = ws.max_row + 1
    ws.append(["v2 per-kit total, this sheet"]
              + [f"={c}{total_row}" for c in "BCDE"])
    v2row = r
    r = ws.max_row + 1
    ws.append(["Increase"] + [f"={c}{v2row}-{c}{v1row}" for c in "BCDE"])
    r = ws.max_row + 1
    ws.append(["Increase, per cent"]
              + [f"=ROUND(({c}{v2row}/{c}{v1row}-1)*100,0)" for c in "BCDE"])
    for rr in range(v1row, r + 1):
        for c in range(2, 6):
            ws.cell(row=rr, column=c).number_format = EUR
    for c in range(2, 6):
        ws.cell(row=r, column=c).number_format = '0"%"'
    ws.append([])
    ws.append(["A v2 kit costs materially more than package v1 estimated, and the reason "
               "is the architecture rather than inflation. v2 buys MODULES where v1 bought "
               "bare silicon (DSN-EEG-003 section 2): no fine-pitch rework, no ADS1299 lead "
               "time, no BGA, but a module carries its own PCB, regulators and headers and "
               "costs several times the chip inside it. The carrier also went to four "
               "layers at 150.0 x 130.0 mm, the case and its foam insert went up several "
               "sizes PAST M-05 to hold the assembled helmet (PKG-EEG-015 section 3.2, "
               "CN-PKG-03), and MP-01, "
               "HM-02, the TPS63020 buck-boost, the 74HC595 driver and the second harness "
               "WH-02 are all new lines that v1 did not carry -- three of them because v1 "
               "had defects that made the kit unbuildable. Some lines went DOWN: ten "
               "electrode cups instead of twelve, and three DIN sockets instead of fifteen."])
    ws.cell(row=ws.max_row, column=1).font = NOTE
    ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    _widths(ws, [46, 15, 15, 15, 15])

    wb.save(path)
    return path


def main():
    print(build(os.path.join(DOCS, f"EEG_kit_BOM_INTERNAL_Rev{REV}_costed.xlsx")))


if __name__ == "__main__":
    main()

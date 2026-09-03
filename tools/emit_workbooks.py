#!/usr/bin/env python3
"""
emit_workbooks.py -- the three spreadsheets, brought up to package v2.

  docs/EEG_kit_BOM_for_bidders_RevC.xlsx        the kit BOM a bidder prices
  (the internal costed BOM is emit_costed_bom.py -- it is a different workbook, not a
   copy of this one with two blank columns, which is what Rev B was)
  docs/EEG_kit_manufacturer_contacts_RevB.xlsx  the contact list, corrected

The kit BOM's carrier sub-BOM sheet is generated from tools/design.py so it can never drift
from the board.  The costed sheet carries the programme's indicative prices and is marked as
such.  The contact list is the package v1 list with the corrections the respondents asked for.

Licence: CC BY-SA 4.0.
"""
from __future__ import annotations
import collections
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HERE = os.path.dirname(os.path.abspath(__file__))
PKG = os.path.dirname(HERE)
V1 = os.path.dirname(PKG)
DOCS = os.path.join(PKG, "docs")
sys.path.insert(0, HERE)

import design as D          # noqa: E402
import pcbgen               # noqa: E402

HEAD = PatternFill("solid", fgColor="1B4F72")
HEADF = Font(color="FFFFFF", bold=True, size=10)
INPUT = PatternFill("solid", fgColor="FFF3C4")
NOTE = Font(italic=True, color="4A5C68", size=9)
BOLD = Font(bold=True, size=10)
THIN = Side(style="thin", color="BFC9D1")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row=1, ncol=None):
    ncol = ncol or ws.max_column
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEAD
        cell.font = HEADF
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 30
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def widths(ws, w):
    for i, x in enumerate(w, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = x


# --------------------------------------------------------------------------- kit BOM
KIT = [
    # group, description, reference part, alternate, qty/kit, substitutable, note
    ("Modules", "ADS1299 8-channel breakout with SPI header and on-board regulators; "
     "DAISY_IN and CLKOUT accessible", "PIEEG-8 module",
     "generic ADS1299 breakout meeting ICD-EEG-006 section 2.1", 2, "No (module class fixed)",
     "Fitted twice per unit. Interface, not brand, is what is specified"),
    ("Modules", "ESP32-S3 development board, 16 MB flash, 8 MB octal PSRAM",
     "Espressif ESP32-S3-DevKitC-1-N16R8", "none", 1, "No",
     "Inserted directly into J6/J7 on 22.86 mm row spacing. GPIO35, 36, 37 and 45 are "
     "not connected on the carrier"),
    ("Modules", "Audio codec module with headphone amplifier, I2S",
     "ES8388 breakout", "WM8960 module (firmware change)", 1, "Yes", ""),
    ("Modules", "USB 2.0 full-speed isolator module, >= 2.5 kV RMS",
     "ADuM4160 isolator module", "none", 1, "No",
     "OPEN: the candidate module presents USB-B and E-24 asks for USB-C. Interim answer "
     "is the WH-09 panel pigtail"),
    ("Modules", "Secure element breakout, I2C", "Adafruit ATECC608 breakout (4314)",
     "any ATECC608B breakout", 1, "No", ""),
    ("Modules", "Li-ion charger with power path and CE pin", "Adafruit bq24074 (4755)",
     "MCP73871 module", 1, "Yes", "Combined charger-plus-gauge assembly is the baseline"),
    ("Modules", "Battery fuel gauge, I2C", "Adafruit MAX17048 (5580)", "BQ27441 module",
     1, "Yes", "May be combined with the charger on one board"),
    ("Modules", "Buck-boost regulator module, VSYS to 5.00 V at 1 A",
     "TPS63020 breakout", "TPS63070 module", 1, "Yes",
     "NEW IN v2. Package v1 had no source for the 3.3 V rail at all (ECO-EEG-002)"),
    ("Modules", "microSD breakout, SDMMC", "generic microSD SDMMC breakout", "none", 1,
     "Yes", "One-bit SDMMC. 70 kB/s needed, about 2 MB/s available"),
    ("Modules", "Electret microphone preamplifier module, fixed gain",
     "MAX4466 class (Adafruit 1063)", "none -- specified by interface, ICD-EEG-006 section 2.9", 1, "Yes",
     "OPEN: the MAX9814 named in v1 is an AGC part and E-14 requires AGC off. Mounts on "
     "MP-01 at J21; the boom carries the bare capsule"),
    ("Modules", "Room microphone module with hardware mute", "to be selected", "none", 1,
     "Yes", "OPEN: no module is known to meet E-15's hardware mute. Fallback is a capsule "
     "with a TS5A3159 analogue switch on an adapter"),
    ("Modules", "8-bit shift register module", "74HC595 breakout", "none", 1, "Yes",
     "Q0..Q7 must be brought out. In v1 they were not, and the contact lights had no "
     "driver at all (ECO-EEG-001)"),
    ("Modules", "Protected 18650 Li-ion cell >= 3000 mAh, UN 38.3",
     "Panasonic NCR18650B protected", "Samsung 35E in a protection holder", 1, "Yes",
     "UN 38.3 test summary required with the first delivery"),
    ("Modules", "microSD card 32 GB high endurance", "SanDisk Industrial",
     "Kingston SDCIT2", 1, "Yes", ""),
    ("Carrier", "EEG-CAR-01 Rev B carrier PCB, FOUR layers, 150.0 x 130.0 mm, ENIG",
     "fabricated from the released Gerber set", "none", 1, "n/a",
     "v1 specified two layers and 130 x 124 mm; see DSN-EEG-003 Rev C section 3"),
    ("Carrier", "EEG-CAR-01 assembly per the carrier sub-BOM sheet",
     "see sheet 'Carrier sub-BOM'", "per sub-BOM alternates", 1, "Per sub-BOM", ""),
    ("Carrier", "Ribbon-jumper set, keyed, per ICD-EEG-006 section 3",
     "2.54 mm IDC ribbon, made up", "none", 1, "Yes", "Maximum 60 mm on the analogue jumper"),
    ("Printed", "HM-01 helmet frame monocoque, PA12 MJF", "MJF PA12, dyed graphite",
     "SLS PA12", 1, "Process only", "About 240 g"),
    ("Printed", "HM-04 electrode assembly body", "MJF PA12", "SLS PA12", 10,
     "Process only", "Eight fitted, two spare"),
    ("Printed", "HM-08 battery hatch (called HM-07 in v1 -- see PARTS-EEG-019)",
     "MJF PA12", "SLS PA12", 1, "Process only", ""),
    ("Printed", "HM-07 boom microphone arm", "MJF PA12 + gooseneck", "none", 1,
     "Process only", ""),
    ("Printed", "MP-01 module mounting plate", "MJF PA12", "none", 1, "Process only",
     "NEW IN v2. 146.0 x 126.0 x 3.0 mm"),
    ("Printed", "POD-P1 enclosure base and lid -- Phase 1 only", "MJF PA12 or FDM PETG",
     "none", 1, "Process only", "163.0 x 143.0 x 58.0 mm external"),
    ("Printed", "HM-02 TPU comfort pads x4 and chin-cup liner", "TPU 85A", "cast", 1,
     "Yes", "Consumable, replaced every turnaround"),
    ("Printed", "HM-09 service key", "MJF PA12", "none", 0, "Process only",
     "ONE PER OPERATOR, deliberately NOT in the participant's kit"),
    ("Printed", "FIT-01 fit-test coupon", "same build as the batch", "none", 0,
     "Process only", "One per print batch, checked before the batch is accepted"),
    ("Bought-in mechanical", "Electrode assembly springs 3-6 N, stainless",
     "per HM-04 drawing", "none", 10, "Yes", ""),
    ("Bought-in mechanical", "Occipital yoke ratchet assembly (POM pawl)",
     "hard-hat style ratchet", "none", 1, "Yes", "2 mm per click, 52-62 cm"),
    ("Bought-in mechanical", "Boom gooseneck 120 mm and mount", "generic", "none", 1,
     "Yes", ""),
    ("Bought-in mechanical", "Chin strap webbing, chin cup and anchors", "generic",
     "none", 1, "Yes", "Removable. The runner says so before it is first mentioned"),
    ("Bought-in mechanical", "M3 x 18 nylon hex standoffs and M3 x 6 screws",
     "Wurth 970200321 class", "none", 4, "Yes", "Carrier to MP-01"),
    ("Bought-in mechanical", "Helmet harness WH-01, 12-way screened, per WH-EEG-008",
     "custom assembly", "none", 1, "Yes",
     "v1 ran the light lines through this cable; v2 splits them out (ECO-EEG-014)"),
    ("Bought-in mechanical", "Contact-light harness WH-02, 10-way, per WH-EEG-008",
     "custom assembly", "none", 1, "Yes", "NEW IN v2"),
    ("Electrodes", "Sintered Ag/AgCl cup electrodes on service bayonet",
     "Wuhan Greentek sintered cup", "Florida Research Instruments / Spes Medica", 10,
     "Yes", "Replaced outright every ~25 sessions"),
    ("Electrodes", "Ag/AgCl ear-clip reference electrodes", "Greentek ear clip", "none",
     2, "Yes", ""),
    ("Electrodes", "EMG snap leads to DIN 42802 plug, 1 m", "Greentek / generic", "none",
     3, "Yes", ""),
    ("Electrodes", "Disposable EMG snap electrodes, pack of 30",
     "Ambu BlueSensor N", "Kendall", 1, "Yes", ""),
    ("Electrodes", "Bicolour contact-light LEDs, two-lead red/green", "generic 3 mm",
     "none", 10, "Yes",
     "NEW IN v2. Two-lead inverse-parallel, driven in two phases at 240 Hz"),
    ("Audio", "Closed-back over-ear headphones, 32 to 64 ohm, 3.5 mm",
     "Audio-Technica ATH-M20x (47 ohm)", "Superlux HD-681", 1, "Yes",
     "A-04 restated to 32-64 ohm; the calibrated level is measured per model"),
    ("Audio", "Electret capsule and windscreen for the boom", "Primo EM272", "generic",
     1, "Yes", "The capsule only; its preamplifier is on MP-01"),
    ("Consumables", "EEG paste 100 g x2, abrasive prep gel 100 g, saline wipes x30, "
     "blunt-tip syringes x4", "Greentek GT20/GT5", "Ten20 / NuPrep", 1, "Yes", ""),
    ("Cables", "USB-C to USB-A 1 m and USB-C to USB-C 1 m", "certified generic", "none",
     1, "Yes", "One of them is the host lead. There is no captive cable in Phase 1"),
    ("Cables", "5 V 2 A USB charger, EU plug, CE", "certified generic", "none", 1, "Yes",
     "The helmet is never worn while the charge cable is connected"),
    ("Case", "Hard-shell IP67 case, internal >= 516 x 390 mm and >= 210 mm deep, of which "
     ">= 185 mm in the base",
     "Peli 1560, unwheeled variant (published internal 518 x 392 x 229 mm)", "none",
     1, "Yes",
     "Sized for an assembled HM-01 standing upright on the halo, beside the POD-P1 bay. "
     "PKG-EEG-015 section 3.2 governs the envelope; Nanuk 915 and Peli 1450 are withdrawn "
     "as about 155 mm deep. Larger than RFQ M-05, see CN-PKG-03. OPEN: no "
     "shell has been bought or measured; the figures are the manufacturer's published ones"),
    ("Case", "Seven-layer PE foam insert, 516 x 390 mm sheet, 25 mm per layer, 175 mm "
     "stack, nine bays", "die-cut or laser-cut to CASE-00 Rev C",
     "pluck foam (Phase 1)", 1, "Yes",
     "OPEN: the CASE-00 Rev C cut file has not been drawn (PKG-EEG-015 section 2.2 and "
     "its open item 1, CN-PKG-02). The two Rev B DXF layers supplied with package v1 are "
     "superseded and must not be cut. Phase 1 packs in pluck foam"),
    ("Case", "Laminated quick-start card and packing photograph; double-wall outer carton",
     "print", "none", 1, "Yes", "IFU-EEG-014 is the artwork"),
    ("Service", "Provisioning, firmware load and TST-EEG-004 Rev C",
     "per TST-EEG-004", "none", 1, "n/a", "About 25 minutes per unit, 20 of them unattended"),
]

QTY = [("Qty @2", 2), ("Qty @10", 10), ("Qty @25", 25), ("Qty @50", 50)]


def kit_bom(path, costed=False):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Kit BOM"
    cols = (["Item", "Group", "Description", "Reference part / module", "Approved alternate",
             "Qty per kit"] + [q[0] for q in QTY]
            + ["Substitutable", "Notes"]
            + (["Indicative unit price EUR", "Extended @10 EUR"] if costed
               else ["Bidder unit price @2", "@10", "@25", "@50", "Bidder notes"]))
    ws.append(cols)
    style_header(ws)
    for i, (grp, desc, ref, alt, q, sub, note) in enumerate(KIT, start=1):
        row = [i, grp, desc, ref, alt, q] + [q * n for _, n in QTY] + [sub, note]
        if costed:
            row += [None, None]
        else:
            row += [None, None, None, None, None]
        ws.append(row)
        for c in range(1, len(cols) + 1):
            ws.cell(row=ws.max_row, column=c).border = BOX
            ws.cell(row=ws.max_row, column=c).alignment = Alignment(
                vertical="top", wrap_text=(c in (3, 4, 5, 11 + len(QTY))))
        if not costed:
            for c in range(len(cols) - 4, len(cols) + 1):
                ws.cell(row=ws.max_row, column=c).fill = INPUT
        else:
            ws.cell(row=ws.max_row, column=len(cols) - 1).fill = INPUT
    widths(ws, [5, 20, 52, 34, 32, 9, 8, 8, 8, 8, 14, 52] + [14] * 6)
    ws.append([])
    ws.append(["", "", "Rev C, 1 September 2026. Four-layer carrier, 150.0 x 130.0 mm, "
                   "in the enlarged POD-P1. Supersedes Rev B entirely. Yellow cells are "
                   "for the bidder; no prices are given in this file by design. "
                   "Nothing in this package has been manufactured or measured."])
    ws.cell(row=ws.max_row, column=3).font = NOTE

    # ---- carrier sub-BOM, generated from design.py
    board = pcbgen.BoardV2()
    board.validate()
    ws2 = wb.create_sheet("Carrier sub-BOM (generated)")
    ws2.append(["Item", "Qty", "Designators", "Value", "Footprint",
                "Manufacturer part number", "Fit", "Notes", "Bidder unit price"])
    style_header(ws2)
    groups = collections.OrderedDict()
    for p in sorted(board.parts, key=lambda p: (p.ref[0], int("".join(
            ch for ch in p.ref if ch.isdigit()) or 0))):
        if p.fpname.startswith("MountingHole"):
            continue
        groups.setdefault((p.value, p.fpname, p.mpn), []).append(p.ref)
    for i, ((val, fpn, mpn), refs) in enumerate(groups.items(), start=1):
        ws2.append([i, len(refs), " ".join(refs), val, fpn, mpn,
                    "DNP" if "DNP" in val else "fit",
                    board.part(refs[0]).descr, None])
        for c in range(1, 10):
            ws2.cell(row=ws2.max_row, column=c).border = BOX
            ws2.cell(row=ws2.max_row, column=c).alignment = Alignment(
                vertical="top", wrap_text=(c in (3, 8)))
        ws2.cell(row=ws2.max_row, column=9).fill = INPUT
    widths(ws2, [5, 6, 40, 26, 34, 32, 6, 60, 14])
    ws2.append([])
    ws2.append(["", "", "GENERATED from package_v2.4/tools/design.py by "
                        "tools/emit_workbooks.py. Do not edit by hand: edit design.py and "
                        "regenerate, so this sheet and the board can never disagree."])
    ws2.cell(row=ws2.max_row, column=3).font = NOTE

    ws3 = wb.create_sheet("Read me")
    for r in [
        ["EEG field kit -- kit bill of materials"],
        ["Revision C, 1 September 2026. Companion to RFQ-EEG-001 Rev E and DSN-EEG-003 Rev C."],
        [],
        ["What changed from Rev B"],
        ["The carrier is four layers and 150.0 x 130.0 mm, not two layers and 130 x 124 mm."],
        ["A buck-boost module is added: package v1 had no source for the 3.3 V rail at all."],
        ["The contact lights are two-lead bicolour LEDs and the 74HC595 outputs now reach "
         "the carrier; in v1 the lights had no driver."],
        ["The helmet harness is two cables, not one."],
        ["MP-01, the module mounting plate, is new, and so are the ribbon jumpers."],
        ["HM-07 is the boom arm and HM-08 is the battery hatch; v1 used HM-07 for both."],
        ["The headphone impedance is restated as 32 to 64 ohm."],
        ["The case and the foam insert carry the two corrections PKG-EEG-015 section 1.2 "
         "asked for and Rev B did not apply: the internal envelope is >= 516 x 390 mm and "
         ">= 210 mm deep, Nanuk 915 and Peli 1450 are withdrawn as too shallow, and the "
         "insert is a seven-layer stack, not the two DXF layers v1 supplied."],
        ["Five items are OPEN and are marked as such: the isolator module's connector, "
         "the boom preamplifier, the room-microphone module, the travel case (no shell "
         "has been bought or measured) and the foam insert (the CASE-00 Rev C cut file "
         "has not been drawn)."],
        [],
        ["Nothing in this package has been manufactured or measured, and no safety "
         "engineer has reviewed the design."],
    ]:
        ws3.append(r)
    ws3.cell(row=1, column=1).font = Font(bold=True, size=13)
    widths(ws3, [110])
    for row in ws3.iter_rows():
        row[0].alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)
    return path


# --------------------------------------------------------------------------- contacts
def contacts(path):
    src = os.path.join(V1, "EEG_kit_manufacturer_contacts.xlsx")
    wb = openpyxl.load_workbook(src)
    ws = wb.worksheets[0]
    # Regulus asked us to correct their certification; the RFQ requests none, so this is
    # informational, but the record should be right.
    fixed = 0
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, str) and "13485" in cell.value:
                cell.value = cell.value.replace("ISO 13485", "ISO 9001")
                fixed += 1
    note = wb.create_sheet("Rev B note")
    for r in [
        ["EEG field kit -- manufacturer contact list, Revision B"],
        ["1 September 2026. Companion to DSN-EEG-003 Rev C and RFQ-EEG-001 Rev E."],
        [],
        [f"Correction applied: {fixed} cell(s) claiming ISO 13485 certification changed to "
         "ISO 9001, at the company's own request. The RFQ requests no certification, so "
         "this is informational."],
        [],
        ["What to send now, and to whom"],
        ["RFQ-EEG-002A -- layout REVIEW (not layout): to any house with a layout desk. "
         "The board is routed; 25 DRC items are open and are named in the DRC report. "
         "Reviewing the routing and closing those items is the scope."],
        ["RFQ-EEG-002B -- fabrication, assembly, printing, harness, provisioning and test, "
         "at 2 / 10 / 25 / 50 units."],
        ["Attach: package_v2.3 in full, or the fabrication subset "
         "kicad/gerber/EEG-CAR-01_RevB_gerber_X2.zip with its layer-map and checksum file."],
        [],
        ["Tell them what changed, because it affects their quote:"],
        ["the carrier is now FOUR layers and 150.0 x 130.0 mm."],
    ]:
        note.append(r)
    note.cell(row=1, column=1).font = Font(bold=True, size=13)
    widths(note, [110])
    for row in note.iter_rows():
        row[0].alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)
    return path


def main():
    print(kit_bom(os.path.join(DOCS, "EEG_kit_BOM_for_bidders_RevC.xlsx")))
    import emit_costed_bom
    print(emit_costed_bom.build(os.path.join(
        DOCS, f"EEG_kit_BOM_INTERNAL_Rev{emit_costed_bom.REV}_costed.xlsx")))
    print(contacts(os.path.join(DOCS, "EEG_kit_manufacturer_contacts_RevB.xlsx")))


if __name__ == "__main__":
    main()
